#!/usr/bin/env python3
"""Export backend paper inventory to an Excel sheet for advisor review."""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from lib.pipeline_layout import DEFAULT_CONFIG_PATH, ensure_pipeline_dirs, load_pipeline_layout

SUPPLEMENT_HINTS = [
    "supplement",
    "additional file",
    "figure s",
    "reporting summary",
    "同行评审",
    "补充",
    "附加",
    "s1.pdf",
    "s2.pdf",
    "s3.pdf",
    "s4.pdf",
    "s5.pdf",
    "s6.pdf",
    "esm",
]

TRAIT_RULES: list[tuple[str, list[str]]] = [
    ("firmness", ["firmness", "crispness", "ripening", "nac18", "expansin"]),
    ("acidity", ["acidity", "malic acid", "malate", "almt"]),
    ("color", ["color", "anthocyanin", "pigmentation", "red-fleshed", "myb10", "carotenoid"]),
    ("disease", ["fire blight", "scab", "mildew", "resistance", "disease"]),
    ("sugar", ["sugar", "sorbitol", "sucrose", "soluble solids", "sweet"]),
    ("maturity", ["maturity", "harvest date", "early ripening", "fruit development"]),
]


def parse_filename(name: str) -> tuple[str, str, str, str]:
    match = re.match(r"^(?P<seq>\d{3})_(?P<slug>.+)_(?P<idx>\d{3})_(?P<title>.+)\.pdf$", name, flags=re.IGNORECASE)
    if match:
        return (
            match.group("seq"),
            match.group("slug"),
            match.group("idx"),
            match.group("title"),
        )
    stem = Path(name).stem
    return ("000", stem, "000", stem)


def is_supplement(title: str) -> bool:
    lowered = title.lower()
    return any(hint in lowered for hint in SUPPLEMENT_HINTS)


def clean_title(title: str) -> str:
    text = title.replace("_", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def detect_trait(text: str) -> str:
    lowered = text.lower()
    for trait, keywords in TRAIT_RULES:
        if any(keyword in lowered for keyword in keywords):
            return trait
    return "unknown"


def choose_main_title(items: list[dict[str, str]]) -> tuple[str, str]:
    primary = [item for item in items if not item["is_supplement"]]
    target = primary or items
    target = sorted(target, key=lambda item: (item["idx"], len(item["title"])))
    chosen = target[0]
    return chosen["title"], chosen["filename"]


def autosize(ws) -> None:
    for column_cells in ws.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        width = min(max(len(v) for v in values) + 2, 60)
        ws.column_dimensions[column_cells[0].column_letter].width = width


def main() -> None:
    parser = argparse.ArgumentParser(description="Export backend paper inventory to advisor review sheet")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG_PATH), help="Pipeline config TOML path")
    parser.add_argument("--papers-dir", help="Override backend paper directory")
    parser.add_argument("--output", help="Override Excel output path")
    args = parser.parse_args()

    layout = load_pipeline_layout(args.config)
    ensure_pipeline_dirs(layout)

    papers_dir = Path(args.papers_dir).resolve() if args.papers_dir else layout.backend_papers_dir
    output = (
        Path(args.output).resolve()
        if args.output
        else layout.reports_dir / "backend_papers_for_advisor_review.xlsx"
    )
    output.parent.mkdir(parents=True, exist_ok=True)

    groups: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for pdf in sorted(papers_dir.glob("*.pdf")):
        seq, slug, idx, title = parse_filename(pdf.name)
        groups[(seq, slug)].append(
            {
                "filename": pdf.name,
                "path": str(pdf),
                "idx": idx,
                "title": clean_title(title),
                "is_supplement": is_supplement(title),
            }
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "papers_for_review"

    headers = [
        "seq",
        "slug",
        "trait_guess",
        "main_title",
        "main_pdf_filename",
        "pdf_file_count",
        "supplement_file_count",
        "all_files",
        "teacher_label",
        "teacher_score",
        "teacher_priority",
        "teacher_notes",
    ]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for (seq, slug), items in sorted(groups.items()):
        main_title, main_filename = choose_main_title(items)
        trait_guess = detect_trait(f"{slug} {main_title}")
        supplement_count = sum(1 for item in items if item["is_supplement"])
        all_files = "\n".join(item["filename"] for item in items)
        ws.append(
            [
                seq,
                slug,
                trait_guess,
                main_title,
                main_filename,
                len(items),
                supplement_count,
                all_files,
                "",
                "",
                "",
                "",
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    autosize(ws)
    wb.save(output)

    print(f"paper_groups={len(groups)}")
    print(f"output={output}")


if __name__ == "__main__":
    main()
